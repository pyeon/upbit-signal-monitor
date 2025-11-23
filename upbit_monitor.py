#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
업비트 실시간 모니터링 시스템 v1.0
거래량 분석 + 기술적 지표 + 호가창 분석을 통한 매매 신호 탐지
"""

import pyupbit
import pandas as pd
import numpy as np
import requests
import time
from datetime import datetime, timedelta
import pytz  # 한국 시간대 사용을 위해 추가
import ta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
import os
warnings.filterwarnings('ignore')

# ============================================
# 한국 시간대 설정
# ============================================
KST = pytz.timezone('Asia/Seoul')

def get_kst_now():
    """한국 시간 반환"""
    return datetime.now(KST)

# ============================================
# 환경변수에서 설정 불러오기
# ============================================

# 텔레그램 봇 설정 (필수)
BOT_TOKEN = os.environ.get('BOT_TOKEN', '')
CHAT_ID = os.environ.get('CHAT_ID', '')

# 모니터링 설정 (선택)
SCAN_INTERVAL = int(os.environ.get('SCAN_INTERVAL', '300'))
VOLUME_THRESHOLD_WATCH = float(os.environ.get('VOLUME_THRESHOLD_WATCH', '1.5'))
VOLUME_THRESHOLD_STRONG = float(os.environ.get('VOLUME_THRESHOLD_STRONG', '2.0'))

# 신호 강도 설정 (선택)
SIGNAL_THRESHOLD_STRONG = int(os.environ.get('SIGNAL_THRESHOLD_STRONG', '7'))
SIGNAL_THRESHOLD_MEDIUM = int(os.environ.get('SIGNAL_THRESHOLD_MEDIUM', '5'))

# 출력 파일 설정 (선택)
EXCEL_FILE = os.environ.get('EXCEL_FILE', 'upbit_signals.xlsx')

# 필수 설정 확인
if not BOT_TOKEN or not CHAT_ID:
    print("❌ 텔레그램 설정이 없습니다!")
    print("\n📝 설정 방법:")
    print("1. GitHub Actions 사용시: Repository Settings → Secrets에 등록")
    print("   - BOT_TOKEN: 텔레그램 봇 토큰")
    print("   - CHAT_ID: 텔레그램 채팅 ID")
    print("\n2. 로컬 실행시: 환경변수로 설정")
    print("   export BOT_TOKEN='your_bot_token'")
    print("   export CHAT_ID='your_chat_id'")
    print("\n3. 또는 config.py 파일 생성:")
    print("   config.example.py를 config.py로 복사 후 값 입력")
    
    # config.py가 있으면 불러오기 시도
    try:
        from config import BOT_TOKEN as CONFIG_BOT_TOKEN
        from config import CHAT_ID as CONFIG_CHAT_ID
        BOT_TOKEN = CONFIG_BOT_TOKEN
        CHAT_ID = CONFIG_CHAT_ID
        print("\n✅ config.py에서 설정을 불러왔습니다.")
    except ImportError:
        print("\n❌ config.py 파일도 없습니다. 프로그램을 종료합니다.")
        exit(1)

# ============================================
# 텔레그램 전송 함수
# ============================================

def send_telegram(message, parse_mode=None):
    """텔레그램 메시지 전송"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {
            "chat_id": CHAT_ID,
            "text": message
        }
        if parse_mode:
            data["parse_mode"] = parse_mode
            
        response = requests.post(url, data=data, timeout=10)
        return response.json()
    except Exception as e:
        print(f"텔레그램 전송 실패: {e}")
        return None

# ============================================
# 거래량 분석 함수
# ============================================

def analyze_volume(coin):
    """거래량 분석 - 4가지 지표"""
    try:
        # 일봉 데이터 (30일)
        df = pyupbit.get_ohlcv(coin, interval="day", count=30)
        if df is None or len(df) < 20:
            return None
        
        current_volume = df['volume'].iloc[-1]
        
        # 1. 거래량 MA 돌파
        volume_ma_20 = df['volume'].rolling(20).mean().iloc[-1]
        volume_ratio = current_volume / volume_ma_20
        
        # 2. 거래량 축적 지수 (7일)
        volume_ma_7 = df['volume'].rolling(7).mean().iloc[-1]
        volume_ma_14 = df['volume'].rolling(14).mean().iloc[-1]
        accumulation_index = ((volume_ma_7 - volume_ma_14) / volume_ma_14) * 100
        
        # 가격 정체 여부 (7일간 가격 변동률)
        price_7d_ago = df['close'].iloc[-8]
        current_price = df['close'].iloc[-1]
        price_change_7d = abs((current_price - price_7d_ago) / price_7d_ago) * 100
        
        # 3. 가격-거래량 괴리도
        price_change_1d = abs((df['close'].iloc[-1] - df['close'].iloc[-2]) / df['close'].iloc[-2]) * 100
        volume_change_1d = ((current_volume - df['volume'].iloc[-2]) / df['volume'].iloc[-2]) * 100
        
        if price_change_1d > 0:
            divergence = volume_change_1d / price_change_1d
        else:
            divergence = 0
        
        return {
            'volume_ratio': volume_ratio,
            'accumulation_index': accumulation_index,
            'price_change_7d': price_change_7d,
            'divergence': divergence,
            'current_volume': current_volume,
            'current_price': current_price
        }
    except Exception as e:
        return None

# ============================================
# 호가창 분석 함수
# ============================================

def analyze_orderbook(coin):
    """호가창 물량 변화 분석"""
    try:
        orderbook = pyupbit.get_orderbook(coin)
        if orderbook is None or not isinstance(orderbook, list) or len(orderbook) == 0:
            return None
        
        orderbook_data = orderbook[0]
        
        if 'orderbook_units' not in orderbook_data:
            return None
        
        units = orderbook_data['orderbook_units']
        
        # 매수/매도 총 물량
        total_bid_size = sum([item.get('bid_size', 0) for item in units])
        total_ask_size = sum([item.get('ask_size', 0) for item in units])
        
        # 매수/매도 비율
        bid_ask_ratio = total_bid_size / total_ask_size if total_ask_size > 0 else 0
        
        # 최상단 호가
        top_bid = units[0].get('bid_size', 0) if len(units) > 0 else 0
        top_ask = units[0].get('ask_size', 0) if len(units) > 0 else 0
        
        return {
            'total_bid': total_bid_size,
            'total_ask': total_ask_size,
            'bid_ask_ratio': bid_ask_ratio,
            'top_bid': top_bid,
            'top_ask': top_ask
        }
    except Exception as e:
        return None

# ============================================
# 기술적 지표 계산 함수
# ============================================

def calculate_indicators(coin):
    """5가지 기술적 지표 계산"""
    try:
        # 일봉 데이터
        df = pyupbit.get_ohlcv(coin, interval="day", count=100)
        if df is None or len(df) < 50:
            return None
        
        # 1. RSI (14)
        rsi = ta.momentum.RSIIndicator(df['close'], window=14).rsi().iloc[-1]
        rsi_signal = "과매도" if rsi < 30 else "과매수" if rsi > 70 else "중립"
        
        # 2. MACD
        macd = ta.trend.MACD(df['close'])
        macd_line = macd.macd().iloc[-1]
        signal_line = macd.macd_signal().iloc[-1]
        macd_hist = macd.macd_diff().iloc[-1]
        macd_signal = "골든크로스" if macd_line > signal_line and macd_hist > 0 else "데드크로스" if macd_line < signal_line and macd_hist < 0 else "중립"
        
        # 3. 볼린저 밴드
        bollinger = ta.volatility.BollingerBands(df['close'])
        bb_high = bollinger.bollinger_hband().iloc[-1]
        bb_low = bollinger.bollinger_lband().iloc[-1]
        current_price = df['close'].iloc[-1]
        
        if current_price >= bb_high:
            bb_signal = "상단터치"
        elif current_price <= bb_low:
            bb_signal = "하단터치"
        else:
            bb_signal = "중립"
        
        # 4. 이동평균선 (5일, 20일)
        ma5 = df['close'].rolling(5).mean().iloc[-1]
        ma20 = df['close'].rolling(20).mean().iloc[-1]
        ma_signal = "상향돌파" if ma5 > ma20 else "하향돌파"
        
        # 5. 거래량 (평균 대비)
        volume_avg = df['volume'].rolling(20).mean().iloc[-1]
        current_volume = df['volume'].iloc[-1]
        volume_percent = (current_volume / volume_avg) * 100
        volume_signal = "급증" if volume_percent > 150 else "정상"
        
        return {
            'rsi': rsi,
            'rsi_signal': rsi_signal,
            'macd_signal': macd_signal,
            'bb_signal': bb_signal,
            'ma_signal': ma_signal,
            'volume_percent': volume_percent,
            'volume_signal': volume_signal,
            'current_price': current_price
        }
    except Exception as e:
        return None

# ============================================
# 신호 강도 판단 함수
# ============================================

def calculate_signal_strength(volume_data, indicators, orderbook_data):
    """9개 지표 기반 신호 강도 계산"""
    score = 0
    signals = []
    
    # 거래량 분석 (4개)
    if volume_data:
        # 1. 거래량 MA 돌파
        if volume_data['volume_ratio'] >= 2.0:
            score += 1
            signals.append("✅ 거래량 MA 돌파")
        
        # 2. 축적 지수
        if volume_data['accumulation_index'] > 20 and volume_data['price_change_7d'] < 5:
            score += 1
            signals.append("✅ 축적 패턴")
        
        # 3. 가격-거래량 괴리
        if volume_data['divergence'] > 10:
            score += 1
            signals.append("✅ 고괴리")
    
    # 호가창 (1개)
    if orderbook_data:
        if orderbook_data['bid_ask_ratio'] > 1.5:
            score += 1
            signals.append("✅ 매수벽 우세")
    
    # 기술적 지표 (5개)
    if indicators:
        if indicators['rsi'] < 30:
            score += 1
            signals.append("✅ RSI 과매도")
        
        if indicators['macd_signal'] == "골든크로스":
            score += 1
            signals.append("✅ MACD 골든크로스")
        
        if indicators['bb_signal'] == "하단터치":
            score += 1
            signals.append("✅ 볼린저 하단")
        
        if indicators['ma_signal'] == "상향돌파":
            score += 1
            signals.append("✅ MA 상향돌파")
        
        if indicators['volume_signal'] == "급증":
            score += 1
            signals.append("✅ 거래량 급증")
    
    return score, signals

# ============================================
# 텔레그램 메시지 포맷팅
# ============================================

def format_telegram_message(coin, score, signals, volume_data, indicators, orderbook_data):
    """텔레그램 메시지 생성"""
    
    # 신호 강도 판단
    if score >= 7:
        emoji = "🔥"
        strength = "강력 매수신호"
        stars = "⭐" * 5
    elif score >= 5:
        emoji = "⚠️"
        strength = "매수 준비신호"
        stars = "⭐" * 3
    else:
        return None
    
    coin_name = coin.replace("KRW-", "")
    
    # 메시지 구성
    message = f"{emoji} [{coin_name}] {strength} {stars}\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"💰 현재가: {volume_data['current_price']:,.0f}원\n\n"
    message += "【 거래량 분석 】\n"
    
    if volume_data['volume_ratio'] >= 1.5:
        ratio_emoji = "✅" if volume_data['volume_ratio'] >= 2.0 else "📊"
        message += f"{ratio_emoji} 거래량 MA 돌파: {volume_data['volume_ratio']:.1f}배\n"
        message += f"   └ 20일 평균 대비 {volume_data['volume_ratio']:.1f}배 ▶ "
        if volume_data['volume_ratio'] >= 2.0:
            message += "강력신호\n"
        else:
            message += "주목\n"
    
    if abs(volume_data['accumulation_index']) > 10:
        message += f"\n📈 축적지수: {volume_data['accumulation_index']:+.1f}%\n"
        if volume_data['accumulation_index'] > 0 and volume_data['price_change_7d'] < 5:
            message += f"   └ 가격 정체({volume_data['price_change_7d']:.1f}%) + 거래량 증가 ▶ 세력 매집 의심\n"
    
    if volume_data['divergence'] > 5:
        message += f"\n⚡ 가격-거래량 괴리: {volume_data['divergence']:.1f}\n"
        message += f"   └ 거래량만 급증 ▶ 큰 움직임 임박\n"
    
    if orderbook_data:
        message += f"\n📊 호가창: 매수/매도 비율 {orderbook_data['bid_ask_ratio']:.2f}\n"
        if orderbook_data['bid_ask_ratio'] > 1.5:
            message += f"   └ 매수벽 우세 ▶ 지지선 형성\n"
    
    message += "\n【 기술적 지표 】\n"
    
    if indicators:
        rsi_emoji = "✅" if indicators['rsi'] < 30 else "📊"
        message += f"{rsi_emoji} RSI: {indicators['rsi']:.1f} → {indicators['rsi_signal']}\n"
        
        macd_emoji = "✅" if indicators['macd_signal'] == '골든크로스' else "📊"
        message += f"{macd_emoji} MACD: {indicators['macd_signal']}\n"
        
        bb_emoji = "✅" if indicators['bb_signal'] == '하단터치' else "📊"
        message += f"{bb_emoji} 볼린저: {indicators['bb_signal']}\n"
        
        ma_emoji = "✅" if indicators['ma_signal'] == '상향돌파' else "📊"
        message += f"{ma_emoji} 이동평균: {indicators['ma_signal']}\n"
        
        vol_emoji = "✅" if indicators['volume_signal'] == '급증' else "📊"
        message += f"{vol_emoji} 거래량: 평균 대비 {indicators['volume_percent']:.0f}% → {indicators['volume_signal']}\n"
    
    message += "\n━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"🎯 종합판단: {score}/9 지표 일치\n"
    message += f"⏰ 발생시각: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}"  # 한국 시간으로 변경
    
    return message

# ============================================
# 엑셀 저장 함수
# ============================================

def save_to_excel(coin, score, volume_data, indicators, orderbook_data):
    """엑셀에 결과 저장"""
    try:
        # 기존 파일 열기 또는 새로 생성
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "실시간 신호"
            
            # 헤더 작성
            headers = ['시간', '코인', '신호강도', '현재가', '거래량비율', '축적지수', 
                      '괴리도', '호가비율', 'RSI', 'MACD', '볼린저', 'MA', '거래량%', '판단']
            ws.append(headers)
            
            # 헤더 스타일
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # 데이터 추가 (한국 시간으로 변경)
        row_data = [
            get_kst_now().strftime('%Y-%m-%d %H:%M:%S'),
            coin.replace('KRW-', ''),
            f"{score}/9",
            volume_data['current_price'] if volume_data else '',
            f"{volume_data['volume_ratio']:.2f}" if volume_data else '',
            f"{volume_data['accumulation_index']:.1f}%" if volume_data else '',
            f"{volume_data['divergence']:.1f}" if volume_data else '',
            f"{orderbook_data['bid_ask_ratio']:.2f}" if orderbook_data else '',
            f"{indicators['rsi']:.1f}" if indicators else '',
            indicators['macd_signal'] if indicators else '',
            indicators['bb_signal'] if indicators else '',
            indicators['ma_signal'] if indicators else '',
            f"{indicators['volume_percent']:.0f}%" if indicators else '',
            "강력매수" if score >= 7 else "매수준비" if score >= 5 else "관심"
        ]
        
        ws.append(row_data)
        
        # 100개 행만 유지
        if ws.max_row > 101:
            ws.delete_rows(2, ws.max_row - 101)
        
        wb.save(EXCEL_FILE)
        print(f"✅ 엑셀 저장 완료: {coin}")
        
    except Exception as e:
        print(f"엑셀 저장 오류: {e}")

# ============================================
# 메인 스캔 함수
# ============================================

def scan_upbit_market():
    """업비트 전체 시장 스캔"""
    print(f"\n{'='*50}")
    print(f"🔍 스캔 시작: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}")  # 한국 시간으로 변경
    print(f"{'='*50}\n")
    
    # 원화 마켓 코인 리스트
    tickers = pyupbit.get_tickers(fiat="KRW")
    print(f"📊 총 {len(tickers)}개 코인 분석 중...\n")
    
    signal_count = 0
    
    for idx, coin in enumerate(tickers, 1):
        try:
            # 진행률 표시
            if idx % 50 == 0:
                print(f"진행률: {idx}/{len(tickers)} ({idx/len(tickers)*100:.1f}%)")
            
            # 1단계: 거래량 분석
            volume_data = analyze_volume(coin)
            if not volume_data or volume_data['volume_ratio'] < VOLUME_THRESHOLD_WATCH:
                continue
            
            print(f"🔎 {coin}: 거래량 {volume_data['volume_ratio']:.1f}배 - 정밀 분석 중...")
            
            # 2단계: 기술적 지표
            indicators = calculate_indicators(coin)
            
            # 3단계: 호가창
            orderbook_data = analyze_orderbook(coin)
            
            # 4단계: 신호 강도 계산
            score, signals = calculate_signal_strength(volume_data, indicators, orderbook_data)
            
            # 5단계: 신호 발송 (5개 이상만)
            if score >= 5:
                signal_count += 1
                
                # 텔레그램 메시지
                message = format_telegram_message(coin, score, signals, volume_data, indicators, orderbook_data)
                if message:
                    send_telegram(message)
                    print(f"✅ 신호 발송: {coin} ({score}/9)")
                
                # 엑셀 저장
                save_to_excel(coin, score, volume_data, indicators, orderbook_data)
            
            # API 제한 방지
            time.sleep(0.1)
            
        except Exception as e:
            print(f"❌ {coin} 분석 오류: {e}")
            continue
    
    print(f"\n{'='*50}")
    print(f"✅ 스캔 완료: 총 {signal_count}개 신호 발견")
    print(f"{'='*50}\n")

# ============================================
# 메인 실행
# ============================================

def main():
    """메인 실행 함수"""
    print("""
    ╔══════════════════════════════════════╗
    ║   업비트 실시간 모니터링 시스템      ║
    ║              v1.0                    ║
    ╚══════════════════════════════════════╝
    """)
    
    # 텔레그램 연결 테스트
    print(f"📱 텔레그램 연결 테스트 중... (Chat ID: {CHAT_ID})")
    test_result = send_telegram("🚀 업비트 모니터링 시스템 시작!")
    
    if test_result and test_result.get('ok'):
        print("✅ 텔레그램 연결 성공!\n")
    else:
        print("❌ 텔레그램 연결 실패!")
        print(f"응답: {test_result}\n")
        print("⚠️  그래도 스캔을 진행합니다...\n")
    
    # 메인 스캔 실행
    try:
        scan_upbit_market()
        
    except KeyboardInterrupt:
        print("\n\n🛑 모니터링 중지됨")
        send_telegram("🛑 업비트 모니터링 시스템 종료")

if __name__ == "__main__":
    main()
