#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
업비트 초단타 급등 조기 감지 시스템 v1.0
5분봉 중심 실시간 모니터링 - 급등 순간 포착
"""

import pyupbit
import pandas as pd
import numpy as np
import requests
import time
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
import os
warnings.filterwarnings('ignore')

# ============================================
# 한국 시간대 설정
# ============================================
KST = pytz.timezone('Asia/Seoul')

def get_kst_now():
    """한국 시간 반환"""
    return datetime.now(KST)

# ============================================
# 환경변수 설정
# ============================================

BOT_TOKEN = os.environ.get('BOT_TOKEN', '')
CHAT_ID = os.environ.get('CHAT_ID', '')

# 🔥 초단타 전용 설정
SCAN_INTERVAL = int(os.environ.get('SCAN_INTERVAL', '120'))  # 2분 스캔
FAST_SCAN_MODE = True  # 빠른 스캔 모드

# 민감도 설정 (더 낮게)
VOLUME_SPIKE_THRESHOLD = float(os.environ.get('VOLUME_SPIKE_THRESHOLD', '1.8'))  # 1.8배면 알림
PRICE_CHANGE_THRESHOLD = float(os.environ.get('PRICE_CHANGE_THRESHOLD', '2.5'))  # 2.5% 상승
CONSECUTIVE_THRESHOLD = int(os.environ.get('CONSECUTIVE_THRESHOLD', '2'))  # 2회 연속

EXCEL_FILE = os.environ.get('EXCEL_FILE', 'upbit_fast_signals.xlsx')

# 설정 확인
if not BOT_TOKEN or not CHAT_ID:
    try:
        from config import BOT_TOKEN as CONFIG_BOT_TOKEN
        from config import CHAT_ID as CONFIG_CHAT_ID
        BOT_TOKEN = CONFIG_BOT_TOKEN
        CHAT_ID = CONFIG_CHAT_ID
    except ImportError:
        print("❌ 텔레그램 설정이 없습니다!")
        exit(1)

# ============================================
# 텔레그램 전송
# ============================================

def send_telegram(message, parse_mode=None):
    """텔레그램 메시지 전송"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {"chat_id": CHAT_ID, "text": message}
        if parse_mode:
            data["parse_mode"] = parse_mode
        response = requests.post(url, data=data, timeout=10)
        return response.json()
    except Exception as e:
        print(f"텔레그램 전송 실패: {e}")
        return None

# ============================================
# 🔥 핵심: 초단타 급등 감지 함수
# ============================================

def detect_price_surge(coin):
    """
    5분봉 기반 급등 조기 감지
    - 거래량 폭발
    - 가격 급등
    - 연속 상승
    """
    try:
        # 5분봉 최근 50개 (약 4시간)
        df = pyupbit.get_ohlcv(coin, interval="minute5", count=50)
        if df is None or len(df) < 20:
            return None
        
        # === 1. 현재 봉 분석 ===
        current_candle = df.iloc[-1]
        current_volume = current_candle['volume']
        current_price = current_candle['close']
        
        # === 2. 거래량 분석 ===
        # 평균 거래량 (직전 10개 봉)
        avg_volume = df['volume'].iloc[-11:-1].mean()
        volume_ratio = current_volume / avg_volume if avg_volume > 0 else 0
        
        # 최근 3개 봉의 거래량 합
        recent_3_volume = df['volume'].iloc[-3:].sum()
        prev_10_volume = df['volume'].iloc[-13:-3].sum()
        volume_acceleration = recent_3_volume / prev_10_volume if prev_10_volume > 0 else 0
        
        # === 3. 가격 분석 ===
        # 현재 봉의 상승률
        candle_change = ((current_candle['close'] - current_candle['open']) / current_candle['open']) * 100
        
        # 5분 전 대비 가격 변화
        price_5m_ago = df['close'].iloc[-2]
        price_change_5m = ((current_price - price_5m_ago) / price_5m_ago) * 100
        
        # 15분 전 대비 가격 변화
        if len(df) >= 4:
            price_15m_ago = df['close'].iloc[-4]
            price_change_15m = ((current_price - price_15m_ago) / price_15m_ago) * 100
        else:
            price_change_15m = 0
        
        # === 4. 연속 상승 분석 ===
        consecutive_green = 0
        for i in range(1, min(6, len(df))):
            if df['close'].iloc[-i] > df['open'].iloc[-i]:  # 양봉
                consecutive_green += 1
            else:
                break
        
        # 연속 거래량 증가
        consecutive_volume = 0
        for i in range(1, min(5, len(df))):
            if df['volume'].iloc[-i] > df['volume'].iloc[-i-1]:
                consecutive_volume += 1
            else:
                break
        
        # === 5. 체결강도 (매수세 분석) ===
        # 최근 5개 봉의 양봉 비율
        recent_5 = df.iloc[-5:]
        green_count = sum(recent_5['close'] > recent_5['open'])
        buying_pressure = green_count / 5
        
        # 고점 돌파 여부
        high_20 = df['high'].iloc[-21:-1].max()
        breaking_high = current_price > high_20
        
        return {
            'volume_ratio': volume_ratio,
            'volume_acceleration': volume_acceleration,
            'candle_change': candle_change,
            'price_change_5m': price_change_5m,
            'price_change_15m': price_change_15m,
            'consecutive_green': consecutive_green,
            'consecutive_volume': consecutive_volume,
            'buying_pressure': buying_pressure,
            'breaking_high': breaking_high,
            'current_price': current_price,
            'current_volume': current_volume
        }
    except Exception as e:
        print(f"급등 감지 오류 ({coin}): {e}")
        return None

# ============================================
# 호가창 실시간 분석
# ============================================

def analyze_orderbook_momentum(coin):
    """호가창 매수/매도 압력 분석"""
    try:
        orderbook = pyupbit.get_orderbook(coin)
        if not orderbook or not isinstance(orderbook, list):
            return None
        
        ob = orderbook[0]
        if 'orderbook_units' not in ob:
            return None
        
        units = ob['orderbook_units']
        
        # 전체 매수/매도 물량
        total_bid = sum([u.get('bid_size', 0) for u in units])
        total_ask = sum([u.get('ask_size', 0) for u in units])
        
        # 상위 3호가 매수/매도
        top3_bid = sum([units[i].get('bid_size', 0) for i in range(min(3, len(units)))])
        top3_ask = sum([units[i].get('ask_size', 0) for i in range(min(3, len(units)))])
        
        # 비율 계산
        bid_ask_ratio = total_bid / total_ask if total_ask > 0 else 0
        top3_ratio = top3_bid / top3_ask if top3_ask > 0 else 0
        
        # 호가창 불균형 (매수벽/매도벽)
        imbalance = (total_bid - total_ask) / (total_bid + total_ask) if (total_bid + total_ask) > 0 else 0
        
        return {
            'bid_ask_ratio': bid_ask_ratio,
            'top3_ratio': top3_ratio,
            'imbalance': imbalance,
            'total_bid': total_bid,
            'total_ask': total_ask
        }
    except Exception as e:
        return None

# ============================================
# 🎯 초단타 신호 판단
# ============================================

def evaluate_fast_signal(surge_data, orderbook_data):
    """
    초단타 신호 강도 평가
    점수 체계: 0-10점
    """
    score = 0
    signals = []
    alert_level = "NORMAL"
    
    if not surge_data:
        return 0, [], "NONE"
    
    # === 거래량 폭발 (0-3점) ===
    if surge_data['volume_ratio'] >= 3.0:
        score += 3
        signals.append("🔥🔥 거래량 3배 폭발")
        alert_level = "CRITICAL"
    elif surge_data['volume_ratio'] >= 2.0:
        score += 2
        signals.append("🔥 거래량 2배 급증")
        alert_level = "HIGH"
    elif surge_data['volume_ratio'] >= 1.5:
        score += 1
        signals.append("⚡ 거래량 1.5배 증가")
    
    # === 가격 급등 (0-3점) ===
    if surge_data['price_change_5m'] >= 5:
        score += 3
        signals.append("🚀🚀 5분 5% 급등")
        alert_level = "CRITICAL"
    elif surge_data['price_change_5m'] >= 3:
        score += 2
        signals.append("🚀 5분 3% 상승")
        if alert_level == "NORMAL":
            alert_level = "HIGH"
    elif surge_data['price_change_5m'] >= 2:
        score += 1
        signals.append("📈 5분 2% 상승")
    
    # === 연속 상승 (0-2점) ===
    if surge_data['consecutive_green'] >= 4:
        score += 2
        signals.append("✅ 4연속 양봉")
    elif surge_data['consecutive_green'] >= 3:
        score += 1
        signals.append("✅ 3연속 양봉")
    
    # === 거래량 가속 (0-1점) ===
    if surge_data['volume_acceleration'] >= 2.0:
        score += 1
        signals.append("⚡ 거래량 가속")
    
    # === 매수세 우위 (0-1점) ===
    if surge_data['buying_pressure'] >= 0.8:
        score += 1
        signals.append("💪 강한 매수세")
    
    # === 고점 돌파 (0-1점) ===
    if surge_data['breaking_high']:
        score += 1
        signals.append("🎯 20봉 고점 돌파")
    
    # === 호가창 매수세 (0-1점) ===
    if orderbook_data:
        if orderbook_data['bid_ask_ratio'] >= 1.8:
            score += 1
            signals.append("💰 호가창 매수벽")
    
    return score, signals, alert_level

# ============================================
# 텔레그램 메시지 포맷
# ============================================

def format_fast_alert(coin, score, signals, surge_data, orderbook_data, alert_level):
    """초단타 알림 메시지"""
    
    # 점수 기준: 6점 이상만 알림
    if score < 6:
        return None
    
    coin_name = coin.replace("KRW-", "")
    
    # 알림 레벨에 따른 이모지
    if alert_level == "CRITICAL":
        emoji = "🚨🔥🔥🔥"
        title = "긴급 급등 알림!"
    elif alert_level == "HIGH":
        emoji = "⚠️🔥"
        title = "급등 감지!"
    else:
        emoji = "📊"
        title = "매수 신호"
    
    message = f"{emoji} [{coin_name}] {title}\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"💰 현재가: {surge_data['current_price']:,.0f}원\n"
    message += f"⭐ 신호강도: {score}/10점\n\n"
    
    # 핵심 지표
    message += "【 실시간 지표 】\n"
    message += f"🔥 거래량: {surge_data['volume_ratio']:.1f}배\n"
    message += f"📈 5분 변화: {surge_data['price_change_5m']:+.2f}%\n"
    
    if surge_data['price_change_15m'] != 0:
        message += f"📈 15분 변화: {surge_data['price_change_15m']:+.2f}%\n"
    
    if surge_data['consecutive_green'] >= 2:
        message += f"✅ 연속 양봉: {surge_data['consecutive_green']}개\n"
    
    if surge_data['buying_pressure'] >= 0.6:
        message += f"💪 매수세: {surge_data['buying_pressure']*100:.0f}%\n"
    
    if orderbook_data:
        message += f"💰 호가 비율: {orderbook_data['bid_ask_ratio']:.2f}\n"
    
    # 신호 목록
    if len(signals) > 0:
        message += "\n【 발생 신호 】\n"
        for sig in signals[:5]:  # 최대 5개만
            message += f"{sig}\n"
    
    message += "\n━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"⏰ {get_kst_now().strftime('%H:%M:%S')}\n"
    message += f"⚡ 즉시 확인 필요!"
    
    return message

# ============================================
# 엑셀 저장
# ============================================

def save_fast_signal(coin, score, surge_data, alert_level):
    """엑셀에 빠르게 저장"""
    try:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "초단타신호"
            
            headers = ['시간', '코인', '레벨', '점수', '현재가', '거래량배수', 
                      '5분변화%', '15분변화%', '연속양봉', '매수세%']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        row = [
            get_kst_now().strftime('%H:%M:%S'),
            coin.replace('KRW-', ''),
            alert_level,
            f"{score}/10",
            surge_data['current_price'],
            f"{surge_data['volume_ratio']:.2f}",
            f"{surge_data['price_change_5m']:+.2f}",
            f"{surge_data['price_change_15m']:+.2f}",
            surge_data['consecutive_green'],
            f"{surge_data['buying_pressure']*100:.0f}"
        ]
        
        ws.append(row)
        
        if ws.max_row > 201:  # 200개만 유지
            ws.delete_rows(2, ws.max_row - 201)
        
        wb.save(EXCEL_FILE)
        
    except Exception as e:
        print(f"엑셀 저장 오류: {e}")

# ============================================
# 메인 스캔
# ============================================

def fast_scan_market():
    """초고속 시장 스캔"""
    print(f"\n⚡ 스캔: {get_kst_now().strftime('%H:%M:%S')}")
    
    tickers = pyupbit.get_tickers(fiat="KRW")
    
    # 빠른 스캔을 위해 시총 상위 코인만 (선택)
    # 또는 전체 스캔
    
    signal_count = 0
    critical_count = 0
    
    for coin in tickers:
        try:
            # 1. 급등 감지
            surge_data = detect_price_surge(coin)
            if not surge_data:
                continue
            
            # 빠른 필터링: 거래량 1.5배 미만은 스킵
            if surge_data['volume_ratio'] < 1.5:
                continue
            
            # 2. 호가창 분석
            orderbook_data = analyze_orderbook_momentum(coin)
            
            # 3. 신호 평가
            score, signals, alert_level = evaluate_fast_signal(surge_data, orderbook_data)
            
            # 4. 알림 발송 (6점 이상)
            if score >= 6:
                signal_count += 1
                if alert_level == "CRITICAL":
                    critical_count += 1
                
                message = format_fast_alert(coin, score, signals, surge_data, orderbook_data, alert_level)
                if message:
                    send_telegram(message)
                    print(f"{'🚨' if alert_level == 'CRITICAL' else '⚠️'} {coin}: {score}/10점")
                
                save_fast_signal(coin, score, surge_data, alert_level)
            
            time.sleep(0.05)  # API 제한
            
        except Exception as e:
            continue
    
    if signal_count > 0:
        print(f"✅ {signal_count}개 신호 (긴급 {critical_count}개)")

# ============================================
# 메인 실행
# ============================================

def main():
    """메인"""
    print("""
    ╔══════════════════════════════════════╗
    ║     업비트 초단타 급등 감지         ║
    ║           v1.0                       ║
    ╚══════════════════════════════════════╝
    """)
    
    send_telegram("⚡ 초단타 급등 감지 시작!")
    
    try:
        fast_scan_market()
        
    except KeyboardInterrupt:
        print("\n🛑 모니터링 종료")
        send_telegram("🛑 초단타 모니터링 종료")

if __name__ == "__main__":
    main()
