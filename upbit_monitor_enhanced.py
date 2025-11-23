#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
업비트 실시간 모니터링 시스템 v2.0 (Enhanced)
일봉 + 단기 시간봉 병행 분석으로 조기 감지 강화
"""

import pyupbit
import pandas as pd
import numpy as np
import requests
import time
from datetime import datetime, timedelta
import pytz
import ta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
import os
warnings.filterwarnings('ignore')

# ============================================
# 한국 시간대 설정
# ============================================
KST = pytz.timezone('Asia/Seoul')

def get_kst_now():
    """한국 시간 반환"""
    return datetime.now(KST)

# ============================================
# 환경변수에서 설정 불러오기
# ============================================

BOT_TOKEN = os.environ.get('BOT_TOKEN', '')
CHAT_ID = os.environ.get('CHAT_ID', '')

# 모니터링 설정
SCAN_INTERVAL = int(os.environ.get('SCAN_INTERVAL', '180'))  # 3분으로 단축
VOLUME_THRESHOLD_WATCH = float(os.environ.get('VOLUME_THRESHOLD_WATCH', '1.3'))  # 더 낮게
VOLUME_THRESHOLD_STRONG = float(os.environ.get('VOLUME_THRESHOLD_STRONG', '2.0'))

# 신호 강도 설정
SIGNAL_THRESHOLD_STRONG = int(os.environ.get('SIGNAL_THRESHOLD_STRONG', '6'))  # 낮춤
SIGNAL_THRESHOLD_MEDIUM = int(os.environ.get('SIGNAL_THRESHOLD_MEDIUM', '4'))  # 낮춤

# 출력 파일 설정
EXCEL_FILE = os.environ.get('EXCEL_FILE', 'upbit_signals_enhanced.xlsx')

# 설정 확인
if not BOT_TOKEN or not CHAT_ID:
    try:
        from config import BOT_TOKEN as CONFIG_BOT_TOKEN
        from config import CHAT_ID as CONFIG_CHAT_ID
        BOT_TOKEN = CONFIG_BOT_TOKEN
        CHAT_ID = CONFIG_CHAT_ID
    except ImportError:
        print("❌ 텔레그램 설정이 없습니다!")
        exit(1)

# ============================================
# 텔레그램 전송 함수
# ============================================

def send_telegram(message, parse_mode=None):
    """텔레그램 메시지 전송"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {
            "chat_id": CHAT_ID,
            "text": message
        }
        if parse_mode:
            data["parse_mode"] = parse_mode
            
        response = requests.post(url, data=data, timeout=10)
        return response.json()
    except Exception as e:
        print(f"텔레그램 전송 실패: {e}")
        return None

# ============================================
# 🆕 단기 시간봉 분석 함수 (핵심 개선)
# ============================================

def analyze_short_term_volume(coin):
    """5분봉, 15분봉 기반 실시간 급등 감지"""
    try:
        # 5분봉 데이터 (최근 100개 = 약 8시간)
        df_5m = pyupbit.get_ohlcv(coin, interval="minute5", count=100)
        # 15분봉 데이터 (최근 100개 = 약 1일)
        df_15m = pyupbit.get_ohlcv(coin, interval="minute15", count=100)
        
        if df_5m is None or df_15m is None or len(df_5m) < 20 or len(df_15m) < 20:
            return None
        
        # === 5분봉 분석 ===
        current_5m_volume = df_5m['volume'].iloc[-1]
        volume_5m_ma_10 = df_5m['volume'].rolling(10).mean().iloc[-1]
        volume_5m_ratio = current_5m_volume / volume_5m_ma_10 if volume_5m_ma_10 > 0 else 0
        
        # 최근 3개 봉의 평균 거래량
        recent_3_volume = df_5m['volume'].iloc[-3:].mean()
        prev_10_volume = df_5m['volume'].iloc[-13:-3].mean()
        volume_surge_ratio = recent_3_volume / prev_10_volume if prev_10_volume > 0 else 0
        
        # 5분봉 가격 변화
        price_change_5m = ((df_5m['close'].iloc[-1] - df_5m['close'].iloc[-4]) / df_5m['close'].iloc[-4]) * 100
        
        # === 15분봉 분석 ===
        current_15m_volume = df_15m['volume'].iloc[-1]
        volume_15m_ma_10 = df_15m['volume'].rolling(10).mean().iloc[-1]
        volume_15m_ratio = current_15m_volume / volume_15m_ma_10 if volume_15m_ma_10 > 0 else 0
        
        # 15분봉 가격 변화
        price_change_15m = ((df_15m['close'].iloc[-1] - df_15m['close'].iloc[-4]) / df_15m['close'].iloc[-4]) * 100
        
        # === 연속 거래량 증가 감지 ===
        consecutive_increase = 0
        for i in range(1, min(5, len(df_5m))):
            if df_5m['volume'].iloc[-i] > df_5m['volume'].iloc[-i-1]:
                consecutive_increase += 1
            else:
                break
        
        # === 체결강도 (간접 계산) ===
        # 양봉/음봉 비율로 매수세 판단
        recent_candles = df_5m.iloc[-10:]
        bullish_count = sum(recent_candles['close'] > recent_candles['open'])
        bullish_ratio = bullish_count / 10
        
        return {
            'volume_5m_ratio': volume_5m_ratio,
            'volume_15m_ratio': volume_15m_ratio,
            'volume_surge_ratio': volume_surge_ratio,
            'price_change_5m': price_change_5m,
            'price_change_15m': price_change_15m,
            'consecutive_increase': consecutive_increase,
            'bullish_ratio': bullish_ratio,
            'current_price': df_5m['close'].iloc[-1]
        }
    except Exception as e:
        print(f"단기 시간봉 분석 오류 ({coin}): {e}")
        return None

# ============================================
# 거래량 분석 함수 (기존 유지)
# ============================================

def analyze_volume(coin):
    """거래량 분석 - 일봉 기반"""
    try:
        df = pyupbit.get_ohlcv(coin, interval="day", count=30)
        if df is None or len(df) < 20:
            return None
        
        current_volume = df['volume'].iloc[-1]
        volume_ma_20 = df['volume'].rolling(20).mean().iloc[-1]
        volume_ratio = current_volume / volume_ma_20
        
        volume_ma_7 = df['volume'].rolling(7).mean().iloc[-1]
        volume_ma_14 = df['volume'].rolling(14).mean().iloc[-1]
        accumulation_index = ((volume_ma_7 - volume_ma_14) / volume_ma_14) * 100
        
        price_7d_ago = df['close'].iloc[-8]
        current_price = df['close'].iloc[-1]
        price_change_7d = abs((current_price - price_7d_ago) / price_7d_ago) * 100
        
        price_change_1d = abs((df['close'].iloc[-1] - df['close'].iloc[-2]) / df['close'].iloc[-2]) * 100
        volume_change_1d = ((current_volume - df['volume'].iloc[-2]) / df['volume'].iloc[-2]) * 100
        
        if price_change_1d > 0:
            divergence = volume_change_1d / price_change_1d
        else:
            divergence = 0
        
        return {
            'volume_ratio': volume_ratio,
            'accumulation_index': accumulation_index,
            'price_change_7d': price_change_7d,
            'divergence': divergence,
            'current_volume': current_volume,
            'current_price': current_price
        }
    except Exception as e:
        return None

# ============================================
# 호가창 분석 함수 (기존 유지)
# ============================================

def analyze_orderbook(coin):
    """호가창 물량 변화 분석"""
    try:
        orderbook = pyupbit.get_orderbook(coin)
        if orderbook is None or not isinstance(orderbook, list) or len(orderbook) == 0:
            return None
        
        orderbook_data = orderbook[0]
        if 'orderbook_units' not in orderbook_data:
            return None
        
        units = orderbook_data['orderbook_units']
        total_bid_size = sum([item.get('bid_size', 0) for item in units])
        total_ask_size = sum([item.get('ask_size', 0) for item in units])
        bid_ask_ratio = total_bid_size / total_ask_size if total_ask_size > 0 else 0
        
        top_bid = units[0].get('bid_size', 0) if len(units) > 0 else 0
        top_ask = units[0].get('ask_size', 0) if len(units) > 0 else 0
        
        return {
            'total_bid': total_bid_size,
            'total_ask': total_ask_size,
            'bid_ask_ratio': bid_ask_ratio,
            'top_bid': top_bid,
            'top_ask': top_ask
        }
    except Exception as e:
        return None

# ============================================
# 기술적 지표 계산 함수 (기존 유지)
# ============================================

def calculate_indicators(coin):
    """5가지 기술적 지표 계산"""
    try:
        df = pyupbit.get_ohlcv(coin, interval="day", count=100)
        if df is None or len(df) < 50:
            return None
        
        rsi = ta.momentum.RSIIndicator(df['close'], window=14).rsi().iloc[-1]
        rsi_signal = "과매도" if rsi < 30 else "과매수" if rsi > 70 else "중립"
        
        macd = ta.trend.MACD(df['close'])
        macd_line = macd.macd().iloc[-1]
        signal_line = macd.macd_signal().iloc[-1]
        macd_hist = macd.macd_diff().iloc[-1]
        macd_signal = "골든크로스" if macd_line > signal_line and macd_hist > 0 else "데드크로스" if macd_line < signal_line and macd_hist < 0 else "중립"
        
        bollinger = ta.volatility.BollingerBands(df['close'])
        bb_high = bollinger.bollinger_hband().iloc[-1]
        bb_low = bollinger.bollinger_lband().iloc[-1]
        current_price = df['close'].iloc[-1]
        
        if current_price >= bb_high:
            bb_signal = "상단터치"
        elif current_price <= bb_low:
            bb_signal = "하단터치"
        else:
            bb_signal = "중립"
        
        ma5 = df['close'].rolling(5).mean().iloc[-1]
        ma20 = df['close'].rolling(20).mean().iloc[-1]
        ma_signal = "상향돌파" if ma5 > ma20 else "하향돌파"
        
        volume_avg = df['volume'].rolling(20).mean().iloc[-1]
        current_volume = df['volume'].iloc[-1]
        volume_percent = (current_volume / volume_avg) * 100
        volume_signal = "급증" if volume_percent > 150 else "정상"
        
        return {
            'rsi': rsi,
            'rsi_signal': rsi_signal,
            'macd_signal': macd_signal,
            'bb_signal': bb_signal,
            'ma_signal': ma_signal,
            'volume_percent': volume_percent,
            'volume_signal': volume_signal,
            'current_price': current_price
        }
    except Exception as e:
        return None

# ============================================
# 🆕 개선된 신호 강도 판단 함수
# ============================================

def calculate_signal_strength(volume_data, indicators, orderbook_data, short_term_data):
    """단기 + 중장기 지표 통합 분석 (최대 14개 지표)"""
    score = 0
    signals = []
    signal_type = "NORMAL"  # EARLY, NORMAL, STRONG
    
    # === 🔥 조기 감지 신호 (단기 시간봉) ===
    if short_term_data:
        # 1. 5분봉 거래량 급증
        if short_term_data['volume_5m_ratio'] >= 2.0:
            score += 2  # 가중치 2배
            signals.append("🔥 5분봉 거래량 폭발")
            signal_type = "EARLY"
        elif short_term_data['volume_5m_ratio'] >= 1.5:
            score += 1
            signals.append("⚡ 5분봉 거래량 증가")
        
        # 2. 연속 거래량 증가
        if short_term_data['consecutive_increase'] >= 3:
            score += 2
            signals.append("🔥 연속 거래량 증가")
            signal_type = "EARLY"
        
        # 3. 급등 진행 중
        if short_term_data['price_change_5m'] > 5:
            score += 2
            signals.append("🚀 5분봉 급등 중")
            signal_type = "EARLY"
        elif short_term_data['price_change_5m'] > 3:
            score += 1
            signals.append("📈 5분봉 상승 중")
        
        # 4. 15분봉 거래량 급증
        if short_term_data['volume_15m_ratio'] >= 2.0:
            score += 1
            signals.append("✅ 15분봉 거래량 돌파")
        
        # 5. 매수세 우위
        if short_term_data['bullish_ratio'] >= 0.7:
            score += 1
            signals.append("✅ 매수세 강함")
    
    # === 일봉 거래량 분석 ===
    if volume_data:
        if volume_data['volume_ratio'] >= 2.0:
            score += 1
            signals.append("✅ 일봉 거래량 MA 돌파")
        
        if volume_data['accumulation_index'] > 20 and volume_data['price_change_7d'] < 5:
            score += 1
            signals.append("✅ 축적 패턴")
        
        if volume_data['divergence'] > 10:
            score += 1
            signals.append("✅ 고괴리")
    
    # === 호가창 ===
    if orderbook_data:
        if orderbook_data['bid_ask_ratio'] > 1.5:
            score += 1
            signals.append("✅ 매수벽 우세")
    
    # === 기술적 지표 ===
    if indicators:
        if indicators['rsi'] < 30:
            score += 1
            signals.append("✅ RSI 과매도")
        
        if indicators['macd_signal'] == "골든크로스":
            score += 1
            signals.append("✅ MACD 골든크로스")
        
        if indicators['bb_signal'] == "하단터치":
            score += 1
            signals.append("✅ 볼린저 하단")
        
        if indicators['ma_signal'] == "상향돌파":
            score += 1
            signals.append("✅ MA 상향돌파")
    
    return score, signals, signal_type

# ============================================
# 🆕 개선된 텔레그램 메시지
# ============================================

def format_telegram_message(coin, score, signals, volume_data, indicators, orderbook_data, short_term_data, signal_type):
    """텔레그램 메시지 생성"""
    
    # 신호 강도 판단
    if signal_type == "EARLY" and score >= 6:
        emoji = "🔥🔥🔥"
        strength = "초단타 급등 감지!"
        stars = "⭐" * 5
    elif score >= 7:
        emoji = "🔥"
        strength = "강력 매수신호"
        stars = "⭐" * 5
    elif score >= 4:
        emoji = "⚠️"
        strength = "매수 준비신호"
        stars = "⭐" * 3
    else:
        return None
    
    coin_name = coin.replace("KRW-", "")
    current_price = short_term_data['current_price'] if short_term_data else volume_data['current_price']
    
    message = f"{emoji} [{coin_name}] {strength} {stars}\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"💰 현재가: {current_price:,.0f}원\n\n"
    
    # 단기 시간봉 정보 (조기 감지 시 강조)
    if short_term_data:
        message += "【 ⚡ 실시간 분석 】\n"
        
        if short_term_data['volume_5m_ratio'] >= 1.5:
            message += f"🔥 5분봉 거래량: {short_term_data['volume_5m_ratio']:.1f}배\n"
        
        if short_term_data['price_change_5m'] > 3:
            message += f"📈 5분 가격변화: +{short_term_data['price_change_5m']:.2f}%\n"
        
        if short_term_data['consecutive_increase'] >= 2:
            message += f"⚡ 연속 증가: {short_term_data['consecutive_increase']}회\n"
        
        if short_term_data['bullish_ratio'] >= 0.6:
            message += f"💪 매수세: {short_term_data['bullish_ratio']*100:.0f}%\n"
        
        message += "\n"
    
    # 기존 정보
    message += "【 거래량 분석 】\n"
    if volume_data and volume_data['volume_ratio'] >= 1.3:
        message += f"📊 일봉 거래량: {volume_data['volume_ratio']:.1f}배\n"
    
    if indicators:
        message += "\n【 기술적 지표 】\n"
        message += f"📊 RSI: {indicators['rsi']:.1f} → {indicators['rsi_signal']}\n"
        message += f"📊 MACD: {indicators['macd_signal']}\n"
    
    message += "\n━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"🎯 종합판단: {score}/14 지표 일치\n"
    message += f"⏰ 발생시각: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    return message

# ============================================
# 엑셀 저장 함수 (개선)
# ============================================

def save_to_excel(coin, score, volume_data, indicators, orderbook_data, short_term_data, signal_type):
    """엑셀에 결과 저장"""
    try:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "실시간 신호"
            
            headers = ['시간', '코인', '신호타입', '신호강도', '현재가', '5분봉거래량', 
                      '가격변화5분', '연속증가', '일봉거래량', 'RSI', '판단']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        current_price = short_term_data['current_price'] if short_term_data else (volume_data['current_price'] if volume_data else '')
        
        row_data = [
            get_kst_now().strftime('%Y-%m-%d %H:%M:%S'),
            coin.replace('KRW-', ''),
            signal_type,
            f"{score}/14",
            current_price,
            f"{short_term_data['volume_5m_ratio']:.2f}" if short_term_data else '',
            f"{short_term_data['price_change_5m']:+.2f}%" if short_term_data else '',
            f"{short_term_data['consecutive_increase']}" if short_term_data else '',
            f"{volume_data['volume_ratio']:.2f}" if volume_data else '',
            f"{indicators['rsi']:.1f}" if indicators else '',
            "🔥조기감지" if signal_type == "EARLY" else "강력매수" if score >= 7 else "매수준비"
        ]
        
        ws.append(row_data)
        
        if ws.max_row > 101:
            ws.delete_rows(2, ws.max_row - 101)
        
        wb.save(EXCEL_FILE)
        
    except Exception as e:
        print(f"엑셀 저장 오류: {e}")

# ============================================
# 🆕 개선된 메인 스캔 함수
# ============================================

def scan_upbit_market():
    """업비트 전체 시장 스캔 (단기 + 중장기 병행)"""
    print(f"\n{'='*50}")
    print(f"🔍 스캔 시작: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*50}\n")
    
    tickers = pyupbit.get_tickers(fiat="KRW")
    print(f"📊 총 {len(tickers)}개 코인 분석 중...\n")
    
    signal_count = 0
    early_detect_count = 0
    
    for idx, coin in enumerate(tickers, 1):
        try:
            if idx % 50 == 0:
                print(f"진행률: {idx}/{len(tickers)} ({idx/len(tickers)*100:.1f}%)")
            
            # 🆕 1단계: 단기 시간봉 먼저 체크 (빠른 감지)
            short_term_data = analyze_short_term_volume(coin)
            
            # 조기 감지 조건: 5분봉 거래량 1.5배 이상 OR 가격 3% 이상 상승
            early_signal = False
            if short_term_data:
                if (short_term_data['volume_5m_ratio'] >= 1.5 or 
                    short_term_data['price_change_5m'] > 3 or
                    short_term_data['consecutive_increase'] >= 3):
                    early_signal = True
                    print(f"⚡ {coin}: 조기 감지! 5분봉 거래량 {short_term_data['volume_5m_ratio']:.1f}배")
            
            # 2단계: 일봉 분석 (기존)
            volume_data = analyze_volume(coin)
            
            # 조기 감지 OR 일봉 조건 충족 시 정밀 분석
            if not early_signal and (not volume_data or volume_data['volume_ratio'] < VOLUME_THRESHOLD_WATCH):
                continue
            
            # 3단계: 기술적 지표 + 호가창
            indicators = calculate_indicators(coin)
            orderbook_data = analyze_orderbook(coin)
            
            # 4단계: 신호 강도 계산
            score, signals, signal_type = calculate_signal_strength(volume_data, indicators, orderbook_data, short_term_data)
            
            # 5단계: 신호 발송 (4개 이상만)
            if score >= 4:
                signal_count += 1
                if signal_type == "EARLY":
                    early_detect_count += 1
                
                message = format_telegram_message(coin, score, signals, volume_data, indicators, orderbook_data, short_term_data, signal_type)
                if message:
                    send_telegram(message)
                    print(f"{'🔥' if signal_type == 'EARLY' else '✅'} 신호 발송: {coin} ({score}/14, {signal_type})")
                
                save_to_excel(coin, score, volume_data, indicators, orderbook_data, short_term_data, signal_type)
            
            time.sleep(0.1)
            
        except Exception as e:
            print(f"❌ {coin} 분석 오류: {e}")
            continue
    
    print(f"\n{'='*50}")
    print(f"✅ 스캔 완료: 총 {signal_count}개 신호 (조기감지 {early_detect_count}개)")
    print(f"{'='*50}\n")

# ============================================
# 메인 실행
# ============================================

def main():
    """메인 실행 함수"""
    print("""
    ╔══════════════════════════════════════╗
    ║   업비트 실시간 모니터링 v2.0       ║
    ║        (단기 시간봉 강화)            ║
    ╚══════════════════════════════════════╝
    """)
    
    print(f"📱 텔레그램 연결 테스트 중...")
    test_result = send_telegram("🚀 업비트 모니터링 v2.0 시작! (단기 시간봉 추가)")
    
    if test_result and test_result.get('ok'):
        print("✅ 텔레그램 연결 성공!\n")
    else:
        print("❌ 텔레그램 연결 실패! 계속 진행...\n")
    
    try:
        scan_upbit_market()
        
    except KeyboardInterrupt:
        print("\n\n🛑 모니터링 중지됨")
        send_telegram("🛑 업비트 모니터링 v2.0 종료")

if __name__ == "__main__":
    main()
